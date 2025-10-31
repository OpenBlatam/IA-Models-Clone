"""
Advanced Export Tools for AI History Comparison System
Herramientas avanzadas de exportación para el sistema de análisis de historial de IA
"""

import asyncio
import json
import csv
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, field
from pathlib import Path
import zipfile
import io
import base64
from enum import Enum
import matplotlib.pyplot as plt
import seaborn as sns
from jinja2 import Template
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import weasyprint
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure matplotlib for better plots
plt.style.use('seaborn-v0_8')
sns.set_palette("husl")

class ExportFormat(Enum):
    """Formatos de exportación disponibles"""
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"
    HTML = "html"
    PNG = "png"
    SVG = "svg"
    ZIP = "zip"

class ExportType(Enum):
    """Tipos de exportación"""
    FULL_REPORT = "full_report"
    STATISTICS = "statistics"
    INSIGHTS = "insights"
    RECOMMENDATIONS = "recommendations"
    PERFORMANCE_DATA = "performance_data"
    ML_RESULTS = "ml_results"
    ALERTS = "alerts"
    TRENDS = "trends"
    CUSTOM = "custom"

@dataclass
class ExportConfig:
    """Configuración de exportación"""
    format: ExportFormat
    export_type: ExportType
    include_charts: bool = True
    include_raw_data: bool = True
    date_range: Optional[tuple] = None
    filters: Dict[str, Any] = field(default_factory=dict)
    custom_template: Optional[str] = None
    output_path: Optional[str] = None
    filename_prefix: str = "ai_history_export"
    compression: bool = False

@dataclass
class ExportResult:
    """Resultado de exportación"""
    success: bool
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    export_time: Optional[float] = None
    error_message: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

class AdvancedExportTools:
    """
    Herramientas avanzadas de exportación para el sistema de análisis de historial de IA
    """
    
    def __init__(
        self,
        output_directory: str = "exports/",
        templates_directory: str = "templates/",
        charts_directory: str = "charts/"
    ):
        self.output_directory = Path(output_directory)
        self.templates_directory = Path(templates_directory)
        self.charts_directory = Path(charts_directory)
        
        # Crear directorios si no existen
        self.output_directory.mkdir(parents=True, exist_ok=True)
        self.templates_directory.mkdir(parents=True, exist_ok=True)
        self.charts_directory.mkdir(parents=True, exist_ok=True)
        
        # Configuración de gráficos
        self.chart_config = {
            "figure_size": (12, 8),
            "dpi": 300,
            "style": "whitegrid",
            "color_palette": "husl"
        }
        
        # Templates HTML
        self.html_templates = self._load_html_templates()
    
    def _load_html_templates(self) -> Dict[str, str]:
        """Cargar templates HTML"""
        templates = {}
        
        # Template para reporte completo
        templates["full_report"] = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Reporte de Análisis de Historial de IA</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }
                .header { text-align: center; margin-bottom: 40px; }
                .section { margin-bottom: 30px; }
                .chart { text-align: center; margin: 20px 0; }
                table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
                th { background-color: #f2f2f2; }
                .metric { display: inline-block; margin: 10px; padding: 15px; background: #f9f9f9; border-radius: 5px; }
                .alert { padding: 10px; margin: 10px 0; border-radius: 5px; }
                .alert.critical { background-color: #f8d7da; border: 1px solid #f5c6cb; }
                .alert.warning { background-color: #fff3cd; border: 1px solid #ffeaa7; }
                .alert.info { background-color: #d1ecf1; border: 1px solid #bee5eb; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 Reporte de Análisis de Historial de IA</h1>
                <p>Generado el {{ timestamp }}</p>
            </div>
            
            <div class="section">
                <h2>📈 Resumen Ejecutivo</h2>
                <div class="metric">
                    <strong>Total de Documentos:</strong> {{ statistics.total_documents }}
                </div>
                <div class="metric">
                    <strong>Calidad Promedio:</strong> {{ "%.2f"|format(statistics.average_quality) }}
                </div>
                <div class="metric">
                    <strong>Alertas Activas:</strong> {{ statistics.active_alerts }}
                </div>
                <div class="metric">
                    <strong>Insights Generados:</strong> {{ statistics.insights_generated }}
                </div>
            </div>
            
            {% if charts %}
            <div class="section">
                <h2>📊 Gráficos de Rendimiento</h2>
                {% for chart in charts %}
                <div class="chart">
                    <h3>{{ chart.title }}</h3>
                    <img src="data:image/png;base64,{{ chart.data }}" alt="{{ chart.title }}" style="max-width: 100%;">
                </div>
                {% endfor %}
            </div>
            {% endif %}
            
            {% if insights %}
            <div class="section">
                <h2>💡 Insights Principales</h2>
                {% for insight in insights %}
                <div class="insight">
                    <h3>{{ insight.title }}</h3>
                    <p><strong>Confianza:</strong> {{ "%.1f"|format(insight.confidence * 100) }}%</p>
                    <p>{{ insight.description }}</p>
                    <ul>
                        {% for rec in insight.recommendations %}
                        <li>{{ rec }}</li>
                        {% endfor %}
                    </ul>
                </div>
                {% endfor %}
            </div>
            {% endif %}
            
            {% if recommendations %}
            <div class="section">
                <h2>🎯 Recomendaciones</h2>
                {% for rec in recommendations %}
                <div class="recommendation">
                    <h3>{{ rec.title }}</h3>
                    <p><strong>Prioridad:</strong> {{ rec.priority }}</p>
                    <p><strong>Mejora Esperada:</strong> {{ "%.1f"|format(rec.expected_improvement * 100) }}%</p>
                    <p>{{ rec.description }}</p>
                </div>
                {% endfor %}
            </div>
            {% endif %}
            
            {% if alerts %}
            <div class="section">
                <h2>🚨 Alertas</h2>
                {% for alert in alerts %}
                <div class="alert {{ alert.level }}">
                    <h4>{{ alert.title }}</h4>
                    <p>{{ alert.message }}</p>
                    <p><small>{{ alert.timestamp }}</small></p>
                </div>
                {% endfor %}
            </div>
            {% endif %}
            
            <div class="section">
                <h2>📋 Datos Detallados</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Métrica</th>
                            <th>Valor</th>
                            <th>Descripción</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for key, value in statistics.items() %}
                        <tr>
                            <td>{{ key }}</td>
                            <td>{{ value }}</td>
                            <td>Métrica del sistema</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </body>
        </html>
        """
        
        return templates
    
    async def export_data(
        self,
        data: Dict[str, Any],
        config: ExportConfig
    ) -> ExportResult:
        """
        Exportar datos según la configuración especificada
        
        Args:
            data: Datos a exportar
            config: Configuración de exportación
            
        Returns:
            Resultado de la exportación
        """
        start_time = datetime.now()
        
        try:
            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{config.filename_prefix}_{config.export_type.value}_{timestamp}"
            
            # Procesar datos según el tipo de exportación
            processed_data = await self._process_export_data(data, config)
            
            # Generar archivo según el formato
            if config.format == ExportFormat.JSON:
                result = await self._export_json(processed_data, filename, config)
            elif config.format == ExportFormat.CSV:
                result = await self._export_csv(processed_data, filename, config)
            elif config.format == ExportFormat.EXCEL:
                result = await self._export_excel(processed_data, filename, config)
            elif config.format == ExportFormat.PDF:
                result = await self._export_pdf(processed_data, filename, config)
            elif config.format == ExportFormat.HTML:
                result = await self._export_html(processed_data, filename, config)
            elif config.format == ExportFormat.PNG:
                result = await self._export_charts(processed_data, filename, config)
            elif config.format == ExportFormat.ZIP:
                result = await self._export_zip(processed_data, filename, config)
            else:
                raise ValueError(f"Formato no soportado: {config.format}")
            
            # Comprimir si se solicita
            if config.compression and result.success:
                result = await self._compress_file(result.file_path)
            
            # Calcular tiempo de exportación
            export_time = (datetime.now() - start_time).total_seconds()
            result.export_time = export_time
            
            # Obtener tamaño del archivo
            if result.file_path and Path(result.file_path).exists():
                result.file_size = Path(result.file_path).stat().st_size
            
            return result
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=str(e),
                export_time=(datetime.now() - start_time).total_seconds()
            )
    
    async def _process_export_data(
        self, 
        data: Dict[str, Any], 
        config: ExportConfig
    ) -> Dict[str, Any]:
        """Procesar datos para exportación"""
        processed = {
            "metadata": {
                "export_timestamp": datetime.now().isoformat(),
                "export_type": config.export_type.value,
                "export_format": config.format.value,
                "filters": config.filters
            }
        }
        
        # Procesar según el tipo de exportación
        if config.export_type == ExportFormat.FULL_REPORT:
            processed.update(data)
        elif config.export_type == ExportFormat.STATISTICS:
            processed["statistics"] = data.get("statistics", {})
        elif config.export_type == ExportFormat.INSIGHTS:
            processed["insights"] = data.get("insights", [])
        elif config.export_type == ExportFormat.RECOMMENDATIONS:
            processed["recommendations"] = data.get("recommendations", [])
        elif config.export_type == ExportFormat.PERFORMANCE_DATA:
            processed["performance_data"] = data.get("performance_data", {})
        elif config.export_type == ExportFormat.ML_RESULTS:
            processed["ml_results"] = data.get("ml_results", {})
        elif config.export_type == ExportFormat.ALERTS:
            processed["alerts"] = data.get("alerts", [])
        elif config.export_type == ExportFormat.TRENDS:
            processed["trends"] = data.get("trends", {})
        
        # Aplicar filtros de fecha si se especifican
        if config.date_range:
            processed = self._apply_date_filters(processed, config.date_range)
        
        # Aplicar filtros adicionales
        if config.filters:
            processed = self._apply_filters(processed, config.filters)
        
        return processed
    
    def _apply_date_filters(self, data: Dict[str, Any], date_range: tuple) -> Dict[str, Any]:
        """Aplicar filtros de fecha"""
        start_date, end_date = date_range
        
        # Filtrar datos que tengan timestamps
        filtered_data = {}
        for key, value in data.items():
            if isinstance(value, list):
                filtered_items = []
                for item in value:
                    if isinstance(item, dict) and "timestamp" in item:
                        item_date = datetime.fromisoformat(item["timestamp"].replace("Z", "+00:00"))
                        if start_date <= item_date <= end_date:
                            filtered_items.append(item)
                    else:
                        filtered_items.append(item)
                filtered_data[key] = filtered_items
            else:
                filtered_data[key] = value
        
        return filtered_data
    
    def _apply_filters(self, data: Dict[str, Any], filters: Dict[str, Any]) -> Dict[str, Any]:
        """Aplicar filtros adicionales"""
        # Implementar filtros específicos según sea necesario
        return data
    
    async def _export_json(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar a JSON"""
        file_path = self.output_directory / f"{filename}.json"
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        return ExportResult(success=True, file_path=str(file_path))
    
    async def _export_csv(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar a CSV"""
        file_path = self.output_directory / f"{filename}.csv"
        
        # Convertir datos a DataFrame
        df = self._data_to_dataframe(data)
        
        # Exportar a CSV
        df.to_csv(file_path, index=False, encoding='utf-8')
        
        return ExportResult(success=True, file_path=str(file_path))
    
    async def _export_excel(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar a Excel con formato avanzado"""
        file_path = self.output_directory / f"{filename}.xlsx"
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Hoja de resumen
            summary_data = {
                'Métrica': list(data.get('statistics', {}).keys()),
                'Valor': list(data.get('statistics', {}).values())
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja de insights
            if 'insights' in data and data['insights']:
                insights_df = pd.DataFrame(data['insights'])
                insights_df.to_excel(writer, sheet_name='Insights', index=False)
            
            # Hoja de recomendaciones
            if 'recommendations' in data and data['recommendations']:
                rec_df = pd.DataFrame(data['recommendations'])
                rec_df.to_excel(writer, sheet_name='Recomendaciones', index=False)
            
            # Hoja de alertas
            if 'alerts' in data and data['alerts']:
                alerts_df = pd.DataFrame(data['alerts'])
                alerts_df.to_excel(writer, sheet_name='Alertas', index=False)
            
            # Aplicar formato
            self._format_excel_sheets(writer)
        
        return ExportResult(success=True, file_path=str(file_path))
    
    def _format_excel_sheets(self, writer):
        """Aplicar formato a las hojas de Excel"""
        workbook = writer.book
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Formatear encabezados
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    async def _export_pdf(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar a PDF con ReportLab"""
        file_path = self.output_directory / f"{filename}.pdf"
        
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        story.append(Paragraph("Reporte de Análisis de Historial de IA", title_style))
        story.append(Spacer(1, 12))
        
        # Fecha de generación
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        story.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", date_style))
        story.append(Spacer(1, 20))
        
        # Resumen estadístico
        if 'statistics' in data:
            story.append(Paragraph("Resumen Estadístico", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            stats_data = [['Métrica', 'Valor']]
            for key, value in data['statistics'].items():
                stats_data.append([key, str(value)])
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 20))
        
        # Insights
        if 'insights' in data and data['insights']:
            story.append(Paragraph("Insights Principales", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            for insight in data['insights'][:5]:  # Máximo 5 insights
                story.append(Paragraph(f"<b>{insight.get('title', 'Sin título')}</b>", styles['Heading3']))
                story.append(Paragraph(insight.get('description', ''), styles['Normal']))
                story.append(Paragraph(f"Confianza: {insight.get('confidence', 0):.1%}", styles['Normal']))
                story.append(Spacer(1, 12))
        
        # Recomendaciones
        if 'recommendations' in data and data['recommendations']:
            story.append(Paragraph("Recomendaciones", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            for rec in data['recommendations'][:5]:  # Máximo 5 recomendaciones
                story.append(Paragraph(f"<b>{rec.get('title', 'Sin título')}</b>", styles['Heading3']))
                story.append(Paragraph(rec.get('description', ''), styles['Normal']))
                story.append(Paragraph(f"Prioridad: {rec.get('priority', 'N/A')}", styles['Normal']))
                story.append(Spacer(1, 12))
        
        # Construir PDF
        doc.build(story)
        
        return ExportResult(success=True, file_path=str(file_path))
    
    async def _export_html(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar a HTML con template"""
        file_path = self.output_directory / f"{filename}.html"
        
        # Usar template personalizado si se especifica
        template_str = config.custom_template or self.html_templates.get("full_report", "")
        
        if not template_str:
            # Template básico si no hay template disponible
            template_str = "<html><body><h1>Datos Exportados</h1><pre>{{ data }}</pre></body></html>"
            data = {"data": json.dumps(data, indent=2, ensure_ascii=False)}
        
        # Renderizar template
        template = Template(template_str)
        html_content = template.render(
            **data,
            timestamp=datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        )
        
        # Escribir archivo
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return ExportResult(success=True, file_path=str(file_path))
    
    async def _export_charts(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar gráficos como imágenes"""
        if not config.include_charts:
            return ExportResult(success=False, error_message="Charts not included in export")
        
        # Generar gráficos
        charts = await self._generate_charts(data)
        
        if not charts:
            return ExportResult(success=False, error_message="No charts generated")
        
        # Guardar gráficos
        chart_files = []
        for i, chart in enumerate(charts):
            chart_path = self.charts_directory / f"{filename}_chart_{i}.png"
            chart.savefig(chart_path, dpi=self.chart_config["dpi"], bbox_inches='tight')
            chart_files.append(str(chart_path))
            plt.close(chart)
        
        return ExportResult(
            success=True, 
            file_path=str(self.charts_directory),
            metadata={"chart_files": chart_files}
        )
    
    async def _export_zip(self, data: Dict[str, Any], filename: str, config: ExportConfig) -> ExportResult:
        """Exportar como archivo ZIP con múltiples formatos"""
        zip_path = self.output_directory / f"{filename}.zip"
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Agregar JSON
            json_data = json.dumps(data, indent=2, ensure_ascii=False, default=str)
            zipf.writestr(f"{filename}.json", json_data)
            
            # Agregar CSV si hay datos tabulares
            df = self._data_to_dataframe(data)
            if not df.empty:
                csv_buffer = io.StringIO()
                df.to_csv(csv_buffer, index=False, encoding='utf-8')
                zipf.writestr(f"{filename}.csv", csv_buffer.getvalue())
            
            # Agregar gráficos si se incluyen
            if config.include_charts:
                charts = await self._generate_charts(data)
                for i, chart in enumerate(charts):
                    chart_buffer = io.BytesIO()
                    chart.savefig(chart_buffer, format='png', dpi=self.chart_config["dpi"], bbox_inches='tight')
                    zipf.writestr(f"charts/chart_{i}.png", chart_buffer.getvalue())
                    plt.close(chart)
        
        return ExportResult(success=True, file_path=str(zip_path))
    
    async def _generate_charts(self, data: Dict[str, Any]) -> List[plt.Figure]:
        """Generar gráficos a partir de los datos"""
        charts = []
        
        # Gráfico de estadísticas
        if 'statistics' in data:
            fig, ax = plt.subplots(figsize=self.chart_config["figure_size"])
            stats = data['statistics']
            
            # Filtrar valores numéricos
            numeric_stats = {k: v for k, v in stats.items() if isinstance(v, (int, float))}
            
            if numeric_stats:
                keys = list(numeric_stats.keys())
                values = list(numeric_stats.values())
                
                bars = ax.bar(keys, values, color=plt.cm.Set3(np.linspace(0, 1, len(keys))))
                ax.set_title('Estadísticas del Sistema', fontsize=16, fontweight='bold')
                ax.set_ylabel('Valor')
                plt.xticks(rotation=45, ha='right')
                
                # Agregar valores en las barras
                for bar, value in zip(bars, values):
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                           f'{value:.2f}', ha='center', va='bottom')
                
                plt.tight_layout()
                charts.append(fig)
        
        # Gráfico de distribución de calidad
        if 'statistics' in data and 'quality_distribution' in data['statistics']:
            fig, ax = plt.subplots(figsize=self.chart_config["figure_size"])
            quality_dist = data['statistics']['quality_distribution']
            
            labels = list(quality_dist.keys())
            sizes = list(quality_dist.values())
            colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99']
            
            wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
            ax.set_title('Distribución de Calidad de Documentos', fontsize=16, fontweight='bold')
            
            # Mejorar la apariencia
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
            
            plt.tight_layout()
            charts.append(fig)
        
        # Gráfico de tendencias si hay datos temporales
        if 'performance_data' in data and 'daily_performance' in data['performance_data']:
            fig, ax = plt.subplots(figsize=self.chart_config["figure_size"])
            daily_perf = data['performance_data']['daily_performance']
            
            days = list(daily_perf.keys())
            quality_scores = [daily_perf[day].get('avg_quality', 0) for day in days]
            
            ax.plot(days, quality_scores, marker='o', linewidth=2, markersize=6)
            ax.set_title('Tendencia de Calidad en el Tiempo', fontsize=16, fontweight='bold')
            ax.set_xlabel('Fecha')
            ax.set_ylabel('Calidad Promedio')
            ax.grid(True, alpha=0.3)
            
            # Formatear fechas en el eje X
            if days:
                ax.set_xticks(range(0, len(days), max(1, len(days)//10)))
                ax.set_xticklabels([days[i] for i in range(0, len(days), max(1, len(days)//10))], rotation=45)
            
            plt.tight_layout()
            charts.append(fig)
        
        return charts
    
    def _data_to_dataframe(self, data: Dict[str, Any]) -> pd.DataFrame:
        """Convertir datos a DataFrame de pandas"""
        rows = []
        
        # Procesar estadísticas
        if 'statistics' in data:
            for key, value in data['statistics'].items():
                rows.append({
                    'Tipo': 'Estadística',
                    'Nombre': key,
                    'Valor': value,
                    'Descripción': f'Estadística del sistema: {key}'
                })
        
        # Procesar insights
        if 'insights' in data:
            for insight in data['insights']:
                rows.append({
                    'Tipo': 'Insight',
                    'Nombre': insight.get('title', 'Sin título'),
                    'Valor': insight.get('confidence', 0),
                    'Descripción': insight.get('description', '')
                })
        
        # Procesar recomendaciones
        if 'recommendations' in data:
            for rec in data['recommendations']:
                rows.append({
                    'Tipo': 'Recomendación',
                    'Nombre': rec.get('title', 'Sin título'),
                    'Valor': rec.get('expected_improvement', 0),
                    'Descripción': rec.get('description', '')
                })
        
        return pd.DataFrame(rows)
    
    async def _compress_file(self, file_path: str) -> ExportResult:
        """Comprimir archivo"""
        try:
            compressed_path = f"{file_path}.gz"
            
            import gzip
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    f_out.writelines(f_in)
            
            # Remover archivo original
            Path(file_path).unlink()
            
            return ExportResult(success=True, file_path=compressed_path)
            
        except Exception as e:
            return ExportResult(success=False, error_message=f"Error comprimiendo archivo: {e}")
    
    async def create_custom_report(
        self,
        data: Dict[str, Any],
        template_path: str,
        output_format: ExportFormat = ExportFormat.HTML
    ) -> ExportResult:
        """Crear reporte personalizado usando template"""
        try:
            # Leer template personalizado
            with open(template_path, 'r', encoding='utf-8') as f:
                template_content = f.read()
            
            # Crear configuración personalizada
            config = ExportConfig(
                format=output_format,
                export_type=ExportType.CUSTOM,
                custom_template=template_content
            )
            
            # Exportar usando template personalizado
            return await self.export_data(data, config)
            
        except Exception as e:
            return ExportResult(success=False, error_message=str(e))
    
    async def batch_export(
        self,
        data: Dict[str, Any],
        formats: List[ExportFormat],
        export_types: List[ExportType]
    ) -> Dict[str, ExportResult]:
        """Exportación en lote a múltiples formatos"""
        results = {}
        
        for export_type in export_types:
            for format_type in formats:
                config = ExportConfig(
                    format=format_type,
                    export_type=export_type
                )
                
                key = f"{export_type.value}_{format_type.value}"
                results[key] = await self.export_data(data, config)
        
        return results
    
    def get_export_history(self) -> List[Dict[str, Any]]:
        """Obtener historial de exportaciones"""
        history = []
        
        for file_path in self.output_directory.glob("*"):
            if file_path.is_file():
                stat = file_path.stat()
                history.append({
                    "filename": file_path.name,
                    "size": stat.st_size,
                    "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
        
        return sorted(history, key=lambda x: x["created"], reverse=True)
    
    async def cleanup_old_exports(self, days: int = 30):
        """Limpiar exportaciones antiguas"""
        cutoff_date = datetime.now() - timedelta(days=days)
        cleaned_count = 0
        
        for file_path in self.output_directory.glob("*"):
            if file_path.is_file():
                file_date = datetime.fromtimestamp(file_path.stat().st_ctime)
                if file_date < cutoff_date:
                    file_path.unlink()
                    cleaned_count += 1
        
        logger.info(f"Cleaned up {cleaned_count} old export files")
        return cleaned_count



























