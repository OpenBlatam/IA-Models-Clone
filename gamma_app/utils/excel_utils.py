"""
Gamma App - Excel Utilities
Advanced Excel processing and manipulation utilities
"""

import io
import base64
from typing import List, Dict, Any, Optional, Union, Tuple
from dataclasses import dataclass
from enum import Enum
import logging
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import xlsxwriter
import xlrd
import xlwt

logger = logging.getLogger(__name__)

class ExcelFormat(Enum):
    """Excel formats"""
    XLSX = "xlsx"
    XLS = "xls"
    CSV = "csv"
    ODS = "ods"

class ChartType(Enum):
    """Chart types"""
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"

@dataclass
class ExcelCell:
    """Excel cell data"""
    row: int
    column: int
    value: Any
    data_type: str
    formula: Optional[str] = None
    style: Optional[Dict[str, Any]] = None

@dataclass
class ExcelSheet:
    """Excel sheet data"""
    name: str
    data: List[List[Any]]
    headers: List[str]
    row_count: int
    column_count: int

@dataclass
class ExcelWorkbook:
    """Excel workbook data"""
    sheets: List[ExcelSheet]
    metadata: Dict[str, Any]
    file_size: int

class ExcelProcessor:
    """Advanced Excel processing class"""
    
    def __init__(self):
        self.supported_formats = ['xlsx', 'xls', 'csv', 'ods']
        self.max_file_size = 50 * 1024 * 1024  # 50MB
    
    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 0
    ) -> pd.DataFrame:
        """Read Excel file into DataFrame"""
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
            else:
                df = pd.read_excel(file_path, header=header_row)
            
            return df
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
    
    def read_excel_sheets(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Read all sheets from Excel file"""
        try:
            sheets = pd.read_excel(file_path, sheet_name=None)
            return sheets
        except Exception as e:
            logger.error(f"Error reading Excel sheets: {e}")
            raise
    
    def write_excel(
        self,
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        output_path: str,
        sheet_name: str = "Sheet1"
    ) -> str:
        """Write data to Excel file"""
        try:
            if isinstance(data, pd.DataFrame):
                data.to_excel(output_path, sheet_name=sheet_name, index=False)
            else:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for sheet_name, df in data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            return output_path
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            raise
    
    def get_workbook_info(self, file_path: str) -> ExcelWorkbook:
        """Get comprehensive workbook information"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            sheets = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get sheet data
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                
                # Get headers (first row)
                headers = data[0] if data else []
                
                sheets.append(ExcelSheet(
                    name=sheet_name,
                    data=data,
                    headers=headers,
                    row_count=sheet.max_row,
                    column_count=sheet.max_column
                ))
            
            # Get metadata
            metadata = {
                'creator': workbook.properties.creator,
                'title': workbook.properties.title,
                'subject': workbook.properties.subject,
                'description': workbook.properties.description,
                'keywords': workbook.properties.keywords,
                'created': workbook.properties.created,
                'modified': workbook.properties.modified,
                'last_modified_by': workbook.properties.lastModifiedBy,
                'version': workbook.properties.version
            }
            
            file_size = Path(file_path).stat().st_size
            
            return ExcelWorkbook(
                sheets=sheets,
                metadata=metadata,
                file_size=file_size
            )
            
        except Exception as e:
            logger.error(f"Error getting workbook info: {e}")
            raise
    
    def create_excel_with_styling(
        self,
        data: Dict[str, pd.DataFrame],
        output_path: str,
        styles: Optional[Dict[str, Any]] = None
    ) -> str:
        """Create Excel file with custom styling"""
        try:
            if styles is None:
                styles = {}
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Apply default styles
                    self._apply_default_styles(worksheet, df)
                    
                    # Apply custom styles if provided
                    if sheet_name in styles:
                        self._apply_custom_styles(worksheet, styles[sheet_name])
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating Excel with styling: {e}")
            raise
    
    def _apply_default_styles(self, worksheet, df: pd.DataFrame):
        """Apply default styles to worksheet"""
        try:
            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in range(1, len(df) + 2):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
                    
        except Exception as e:
            logger.error(f"Error applying default styles: {e}")
    
    def _apply_custom_styles(self, worksheet, styles: Dict[str, Any]):
        """Apply custom styles to worksheet"""
        try:
            # Apply cell styles
            if 'cells' in styles:
                for cell_style in styles['cells']:
                    cell = worksheet.cell(
                        row=cell_style['row'],
                        column=cell_style['column']
                    )
                    
                    if 'font' in cell_style:
                        font_style = cell_style['font']
                        cell.font = Font(
                            bold=font_style.get('bold', False),
                            italic=font_style.get('italic', False),
                            color=font_style.get('color', '000000'),
                            size=font_style.get('size', 11)
                        )
                    
                    if 'fill' in cell_style:
                        fill_style = cell_style['fill']
                        cell.fill = PatternFill(
                            start_color=fill_style.get('color', 'FFFFFF'),
                            end_color=fill_style.get('color', 'FFFFFF'),
                            fill_type='solid'
                        )
                    
                    if 'alignment' in cell_style:
                        align_style = cell_style['alignment']
                        cell.alignment = Alignment(
                            horizontal=align_style.get('horizontal', 'left'),
                            vertical=align_style.get('vertical', 'center')
                        )
            
            # Apply row styles
            if 'rows' in styles:
                for row_style in styles['rows']:
                    row_num = row_style['row']
                    
                    if 'height' in row_style:
                        worksheet.row_dimensions[row_num].height = row_style['height']
                    
                    if 'fill' in row_style:
                        fill_color = row_style['fill']['color']
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_num, column=col)
                            cell.fill = PatternFill(
                                start_color=fill_color,
                                end_color=fill_color,
                                fill_type='solid'
                            )
            
            # Apply column styles
            if 'columns' in styles:
                for col_style in styles['columns']:
                    col_letter = get_column_letter(col_style['column'])
                    
                    if 'width' in col_style:
                        worksheet.column_dimensions[col_letter].width = col_style['width']
                    
                    if 'fill' in col_style:
                        fill_color = col_style['fill']['color']
                        for row in range(1, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_style['column'])
                            cell.fill = PatternFill(
                                start_color=fill_color,
                                end_color=fill_color,
                                fill_type='solid'
                            )
                            
        except Exception as e:
            logger.error(f"Error applying custom styles: {e}")
    
    def add_chart(
        self,
        file_path: str,
        output_path: str,
        sheet_name: str,
        chart_type: ChartType,
        data_range: str,
        title: str = "Chart",
        position: str = "E2"
    ) -> str:
        """Add chart to Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[sheet_name]
            
            # Create chart based on type
            if chart_type == ChartType.BAR:
                chart = BarChart()
            elif chart_type == ChartType.LINE:
                chart = LineChart()
            elif chart_type == ChartType.PIE:
                chart = PieChart()
            else:
                chart = BarChart()  # Default to bar chart
            
            # Set chart properties
            chart.title = title
            chart.style = 13
            
            # Add data to chart
            data = Reference(worksheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            
            # Add chart to worksheet
            worksheet.add_chart(chart, position)
            
            # Save workbook
            workbook.save(output_path)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error adding chart: {e}")
            raise
    
    def filter_data(
        self,
        file_path: str,
        output_path: str,
        sheet_name: str,
        filters: Dict[str, Any]
    ) -> str:
        """Filter data in Excel file"""
        try:
            # Read data
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Apply filters
            for column, filter_value in filters.items():
                if column in df.columns:
                    if isinstance(filter_value, list):
                        df = df[df[column].isin(filter_value)]
                    else:
                        df = df[df[column] == filter_value]
            
            # Write filtered data
            df.to_excel(output_path, sheet_name=sheet_name, index=False)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error filtering data: {e}")
            raise
    
    def sort_data(
        self,
        file_path: str,
        output_path: str,
        sheet_name: str,
        sort_columns: List[str],
        ascending: List[bool] = None
    ) -> str:
        """Sort data in Excel file"""
        try:
            # Read data
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Set default ascending values
            if ascending is None:
                ascending = [True] * len(sort_columns)
            
            # Sort data
            df = df.sort_values(by=sort_columns, ascending=ascending)
            
            # Write sorted data
            df.to_excel(output_path, sheet_name=sheet_name, index=False)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error sorting data: {e}")
            raise
    
    def pivot_table(
        self,
        file_path: str,
        output_path: str,
        sheet_name: str,
        index: List[str],
        columns: List[str],
        values: List[str],
        aggfunc: str = "sum"
    ) -> str:
        """Create pivot table"""
        try:
            # Read data
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Create pivot table
            pivot = df.pivot_table(
                index=index,
                columns=columns,
                values=values,
                aggfunc=aggfunc,
                fill_value=0
            )
            
            # Write pivot table
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                pivot.to_excel(writer, sheet_name=f"{sheet_name}_pivot")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating pivot table: {e}")
            raise
    
    def merge_sheets(
        self,
        file_path: str,
        output_path: str,
        sheet_names: List[str],
        merge_key: str
    ) -> str:
        """Merge multiple sheets"""
        try:
            # Read all sheets
            sheets = pd.read_excel(file_path, sheet_name=sheet_names)
            
            # Merge sheets
            merged_df = sheets[sheet_names[0]]
            for sheet_name in sheet_names[1:]:
                merged_df = pd.merge(
                    merged_df,
                    sheets[sheet_name],
                    on=merge_key,
                    how='outer'
                )
            
            # Write merged data
            merged_df.to_excel(output_path, sheet_name="merged", index=False)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error merging sheets: {e}")
            raise
    
    def calculate_formulas(
        self,
        file_path: str,
        output_path: str,
        sheet_name: str,
        formulas: Dict[str, str]
    ) -> str:
        """Add calculated formulas to Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[sheet_name]
            
            # Add formulas
            for cell_address, formula in formulas.items():
                worksheet[cell_address] = formula
            
            # Save workbook
            workbook.save(output_path)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error adding formulas: {e}")
            raise
    
    def convert_to_csv(
        self,
        file_path: str,
        output_path: str,
        sheet_name: str = None
    ) -> str:
        """Convert Excel to CSV"""
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            df.to_csv(output_path, index=False)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error converting to CSV: {e}")
            raise
    
    def convert_from_csv(
        self,
        csv_path: str,
        output_path: str,
        sheet_name: str = "Sheet1"
    ) -> str:
        """Convert CSV to Excel"""
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(output_path, sheet_name=sheet_name, index=False)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error converting from CSV: {e}")
            raise
    
    def validate_excel(self, file_path: str) -> Dict[str, Any]:
        """Validate Excel file"""
        try:
            validation_result = {
                'valid': True,
                'errors': [],
                'warnings': [],
                'info': {}
            }
            
            # Check if file exists
            if not Path(file_path).exists():
                validation_result['valid'] = False
                validation_result['errors'].append("File does not exist")
                return validation_result
            
            # Check file size
            file_size = Path(file_path).stat().st_size
            if file_size > self.max_file_size:
                validation_result['warnings'].append(f"File size ({file_size} bytes) exceeds maximum ({self.max_file_size} bytes)")
            
            # Check file extension
            file_ext = Path(file_path).suffix.lower()
            if file_ext not in ['.xlsx', '.xls', '.csv']:
                validation_result['warnings'].append(f"File extension {file_ext} may not be supported")
            
            # Try to open file
            try:
                if file_ext == '.csv':
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)
                
                validation_result['info']['row_count'] = len(df)
                validation_result['info']['column_count'] = len(df.columns)
                validation_result['info']['columns'] = list(df.columns)
                
            except Exception as e:
                validation_result['valid'] = False
                validation_result['errors'].append(f"Cannot open file: {str(e)}")
            
            return validation_result
            
        except Exception as e:
            logger.error(f"Error validating Excel file: {e}")
            return {
                'valid': False,
                'errors': [f"Validation error: {str(e)}"],
                'warnings': [],
                'info': {}
            }
    
    def get_data_summary(self, file_path: str, sheet_name: str = None) -> Dict[str, Any]:
        """Get data summary from Excel file"""
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            summary = {
                'shape': df.shape,
                'columns': list(df.columns),
                'dtypes': df.dtypes.to_dict(),
                'null_counts': df.isnull().sum().to_dict(),
                'numeric_summary': df.describe().to_dict() if len(df.select_dtypes(include=['number']).columns) > 0 else {},
                'memory_usage': df.memory_usage(deep=True).sum()
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Error getting data summary: {e}")
            return {}

# Global Excel processor instance
excel_processor = ExcelProcessor()

def read_excel_file(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """Read Excel file using global processor"""
    return excel_processor.read_excel(file_path, sheet_name)

def write_excel_file(data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], output_path: str, sheet_name: str = "Sheet1") -> str:
    """Write Excel file using global processor"""
    return excel_processor.write_excel(data, output_path, sheet_name)

def get_excel_info(file_path: str) -> ExcelWorkbook:
    """Get Excel info using global processor"""
    return excel_processor.get_workbook_info(file_path)

def create_styled_excel(data: Dict[str, pd.DataFrame], output_path: str, styles: Dict[str, Any] = None) -> str:
    """Create styled Excel using global processor"""
    return excel_processor.create_excel_with_styling(data, output_path, styles)

def add_excel_chart(file_path: str, output_path: str, sheet_name: str, chart_type: ChartType, data_range: str, title: str = "Chart") -> str:
    """Add chart to Excel using global processor"""
    return excel_processor.add_chart(file_path, output_path, sheet_name, chart_type, data_range, title)

def filter_excel_data(file_path: str, output_path: str, sheet_name: str, filters: Dict[str, Any]) -> str:
    """Filter Excel data using global processor"""
    return excel_processor.filter_data(file_path, output_path, sheet_name, filters)

def sort_excel_data(file_path: str, output_path: str, sheet_name: str, sort_columns: List[str], ascending: List[bool] = None) -> str:
    """Sort Excel data using global processor"""
    return excel_processor.sort_data(file_path, output_path, sheet_name, sort_columns, ascending)

def create_pivot_table(file_path: str, output_path: str, sheet_name: str, index: List[str], columns: List[str], values: List[str], aggfunc: str = "sum") -> str:
    """Create pivot table using global processor"""
    return excel_processor.pivot_table(file_path, output_path, sheet_name, index, columns, values, aggfunc)

def convert_excel_to_csv(file_path: str, output_path: str, sheet_name: str = None) -> str:
    """Convert Excel to CSV using global processor"""
    return excel_processor.convert_to_csv(file_path, output_path, sheet_name)

def convert_csv_to_excel(csv_path: str, output_path: str, sheet_name: str = "Sheet1") -> str:
    """Convert CSV to Excel using global processor"""
    return excel_processor.convert_from_csv(csv_path, output_path, sheet_name)

























